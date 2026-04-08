import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

# Conexão com o SQL Server
conn = pyodbc.connect(
    'DRIVER={SQL Server};'
    'SERVER=BRUNO\\SQLEXPRESS;'
    'DATABASE=Bd_ESTOQUE;'
    'Trusted_Connection=yes;'
)

# Lendo os dados
query = """
    SELECT 
        p.nome AS Produto,
        c.nome AS Categoria,
        f.nome AS Fornecedor,
        p.preco_unitario AS Preco,
        p.quantidade_estoque AS Estoque,
        p.estoque_minimo AS Minimo
    FROM Produtos p
    INNER JOIN Categorias c ON p.id_categoria = c.id_categoria
    INNER JOIN Fornecedores f ON p.id_fornecedor = f.id_fornecedor
"""

df = pd.read_sql(query, conn)
conn.close()

# Coluna de status
df['Status'] = df.apply(
    lambda row: '⚠️ Repor' if row['Estoque'] <= row['Minimo'] else '✅ OK', axis=1
)

# Nome do arquivo com data
data_hoje = datetime.now().strftime('%d-%m-%Y')
nome_arquivo = f'relatorio_estoque_{data_hoje}.xlsx'

# Exportar pro Excel
df.to_excel(nome_arquivo, index=False, sheet_name='Estoque')

# Formatação
wb = load_workbook(nome_arquivo)
ws = wb['Estoque']

# Cabeçalho em azul
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Linhas com estoque baixo em vermelho
red_fill = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

for row in ws.iter_rows(min_row=2):
    status = row[6].value
    for cell in row:
        if status == '⚠️ Repor':
            cell.fill = red_fill
        else:
            cell.fill = green_fill

# Ajustar largura das colunas
for col in ws.columns:
    max_length = max(len(str(cell.value or '')) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_length + 4

wb.save(nome_arquivo)
print(f"✅ Relatório gerado com sucesso: {nome_arquivo}")